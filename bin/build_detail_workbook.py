#!/usr/bin/env python3
"""Build a per-patient Excel workbook with per-cell/per-variant detail.

Companion to the SNV HTML report: the report heatmaps show clustering
and flag status at a glance but cannot show the exact WBC DP/alt-read
count behind a given flagged variant, or a given cell raw DP for a
variant it does not carry. This workbook makes those numbers directly
filterable/lookup-able without clicking through the report.
"""

import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MIN_ALT_READS = 4

COLORS = {
    "good": ("E6F5E6", "0CA30C"),
    "warning": ("FBF1DE", "C98500"),
    "critical": ("FBE9E7", "D03B3B"),
    "info": ("EAF2FC", "2A78D6"),
    "neutral": ("F2F2F2", "666666"),
}

CTC_STATE_STYLE = {
    "present_pass": "good",
    "present_lowconf": "warning",
    "filtered_out": "neutral",
    "absent": "neutral",
    "not_covered": "neutral",
}
CTC_STATE_LABEL = {
    "present_pass": "present (PASS)",
    "present_lowconf": "present (LOWCONF)",
    "filtered_out": "filtered out (germline/pop-AF)",
    "absent": "absent (0/0)",
    "not_covered": "not covered",
}
WBC_TIER_STYLE = {
    "clean": "good",
    "not_covered": "neutral",
    "weak_signal": "warning",
    "strong_signal": "critical",
    "not_available": "neutral",
}

LONG_COLUMNS = [
    "ctc_id", "cell_id", "sample_name", "var_id", "CHROM", "POS", "REF", "ALT",
    "QUAL", "FILTER", "GT", "DP", "AD", "refread", "altread",
    "gene_symbol", "consequence", "impact", "hgvsc", "hgvsp",
    "existing_variation", "max_af", "max_af_pops",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def natural_sort_key(s):
    return [int(t) if t.isdigit() else t for t in re.split(r"(\d+)", str(s))]


def parse_sample(format_field, sample_field):
    keys = format_field.split(":") if format_field else []
    values = sample_field.split(":") if sample_field else []
    return {k: values[i] if i < len(values) else "" for i, k in enumerate(keys)}


def parse_ad(ad):
    if not ad or ad == ".":
        return None, None
    parts = ad.split(",")
    try:
        ref = int(parts[0]) if parts[0] not in ("", ".") else None
        alt = int(parts[1]) if len(parts) >= 2 and parts[1] not in ("", ".") else None
        return ref, alt
    except ValueError:
        return None, None


def load_raw_records(vcf_path, wanted_keys):
    """Return chrom/pos/ref/alt keyed dict of gt/dp/alt for wanted_keys found in vcf_path."""
    found = {}
    if not vcf_path or not wanted_keys or not Path(vcf_path).exists():
        return found
    with open(vcf_path, newline="") as handle:
        for line in handle:
            if line.startswith("#"):
                continue
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 10:
                continue
            key = (fields[0], fields[1], fields[3], fields[4])
            if key not in wanted_keys:
                continue
            sample = parse_sample(fields[8], fields[9])
            _, alt = parse_ad(sample.get("AD", ""))
            found[key] = {
                "gt": sample.get("GT", ""),
                "dp": sample.get("DP", ""),
                "alt": alt if alt is not None else "",
            }
    return found


def ctc_state(rec):
    if rec is None:
        return "not_covered"
    gt = rec["gt"]
    if gt in ("./.", ".|.", ""):
        return "not_covered"
    if gt in ("0/0", "0|0"):
        return "absent"
    return "filtered_out"


def wbc_tier(rec):
    if rec is None:
        return "not_available", "", "", ""
    gt = rec["gt"]
    if gt in ("0/0", "0|0"):
        return "clean", gt, rec["dp"], rec["alt"]
    if gt in ("./.", ".|.", ""):
        return "not_covered", gt, rec["dp"], rec["alt"]
    alt = rec["alt"] if isinstance(rec["alt"], int) else 0
    return ("strong_signal" if alt >= MIN_ALT_READS else "weak_signal"), gt, rec["dp"], rec["alt"]


def load_long_table(path):
    rows = []
    with open(path, newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            rows.append(row)
    return rows


def map_split_vcfs_by_cell(paths):
    mapping = {}
    for p in paths:
        cell_id = Path(p).name.replace(".monovar.split.vcf", "")
        mapping[cell_id] = p
    return mapping


def style_cell(cell, palette_key, bold=False):
    if palette_key not in COLORS:
        return
    fill_hex, font_hex = COLORS[palette_key]
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.font = Font(color=font_hex, bold=bold)


def write_header(ws, headers):
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"


def autosize(ws, headers, max_width=40):
    for col_idx, name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width = max(len(str(name)) + 2, 10)
        ws.column_dimensions[letter].width = min(width, max_width)


def build_all_variants_sheet(wb, rows, wbc_raw):
    ws = wb.create_sheet("All final variants")
    headers = LONG_COLUMNS + ["VAF", "WBC_tier", "WBC_GT", "WBC_DP", "WBC_alt"]
    write_header(ws, headers)

    for r_idx, row in enumerate(rows, start=2):
        key = (row.get("CHROM", ""), row.get("POS", ""), row.get("REF", ""), row.get("ALT", ""))
        ref_n, alt_n = row.get("refread"), row.get("altread")
        try:
            ref_n, alt_n = float(ref_n), float(alt_n)
            vaf = alt_n / (ref_n + alt_n) if (ref_n + alt_n) > 0 else ""
        except (TypeError, ValueError):
            vaf = ""
        tier, gt, dp, alt = wbc_tier(wbc_raw.get(key))

        for c_idx, col in enumerate(LONG_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(col, ""))
            if col == "FILTER":
                style_cell(cell, "good" if row.get("FILTER") == "PASS" else "warning")
        ws.cell(row=r_idx, column=len(LONG_COLUMNS) + 1, value=round(vaf, 3) if isinstance(vaf, float) else vaf)
        tier_cell = ws.cell(row=r_idx, column=len(LONG_COLUMNS) + 2, value=tier)
        style_cell(tier_cell, WBC_TIER_STYLE.get(tier, "neutral"))
        ws.cell(row=r_idx, column=len(LONG_COLUMNS) + 3, value=gt)
        ws.cell(row=r_idx, column=len(LONG_COLUMNS) + 4, value=dp)
        ws.cell(row=r_idx, column=len(LONG_COLUMNS) + 5, value=alt)

    autosize(ws, headers)


def compute_shared_variant_states(rows, ctc_ids, split_vcf_by_cell, wbc_raw):
    """One entry per variant shared by >=2 CTCs: metadata, per-CTC state/DP/alt, WBC tier.

    Shared by the Excel "Shared variants detail" sheet and the SNV
    report's shared-variant heatmap, so the two stay consistent instead
    of drifting from two separate implementations.
    """
    var_meta = {}
    var_cells = {}
    for row in rows:
        var_id = row["var_id"]
        var_cells.setdefault(var_id, {})[row["ctc_id"]] = row
        if var_id not in var_meta:
            var_meta[var_id] = row

    shared = {v: cells for v, cells in var_cells.items() if len(cells) >= 2}

    needed_by_ctc = {c: set() for c in ctc_ids}
    for var_id, cells in shared.items():
        m = var_meta[var_id]
        key = (m["CHROM"], m["POS"], m["REF"], m["ALT"])
        for ctc in ctc_ids:
            if ctc not in cells:
                needed_by_ctc[ctc].add(key)
    raw_by_ctc = {c: load_raw_records(split_vcf_by_cell.get(c), needed_by_ctc[c]) for c in ctc_ids}

    ordered = sorted(shared, key=lambda v: -len(shared[v]))
    computed = []
    for var_id in ordered:
        m = var_meta[var_id]
        key = (m["CHROM"], m["POS"], m["REF"], m["ALT"])
        cells = shared[var_id]

        states = {}
        for ctc in ctc_ids:
            if ctc in cells:
                final_row = cells[ctc]
                state = "present_pass" if final_row.get("FILTER") == "PASS" else "present_lowconf"
                dp, alt = final_row.get("DP", ""), final_row.get("altread", "")
            else:
                rec = raw_by_ctc[ctc].get(key)
                state = ctc_state(rec)
                dp = rec["dp"] if rec else ""
                alt = rec["alt"] if rec else ""
            states[ctc] = {"state": state, "dp": dp, "alt": alt}

        tier, gt, dp, alt = wbc_tier(wbc_raw.get(key))
        computed.append({
            "var_id": var_id,
            "gene_symbol": m.get("gene_symbol", ""),
            "CHROM": m["CHROM"],
            "POS": m["POS"],
            "REF": m["REF"],
            "ALT": m["ALT"],
            "consequence": m.get("consequence", ""),
            "impact": m.get("impact", ""),
            "shared_n": len(cells),
            "states": states,
            "wbc": {"tier": tier, "gt": gt, "dp": dp, "alt": alt},
        })
    return computed


def build_shared_variants_sheet(wb, computed_rows, ctc_ids):
    ws = wb.create_sheet("Shared variants detail")
    base_headers = ["var_id", "gene_symbol", "CHROM", "POS", "REF", "ALT", "consequence", "impact", "shared_n"]
    ctc_headers = [h for c in ctc_ids for h in (f"{c}_state", f"{c}_DP", f"{c}_alt")]
    wbc_headers = ["WBC_tier", "WBC_GT", "WBC_DP", "WBC_alt"]
    headers = base_headers + ctc_headers + wbc_headers
    write_header(ws, headers)

    for r_idx, entry in enumerate(computed_rows, start=2):
        col = 1
        for val in (entry["var_id"], entry["gene_symbol"], entry["CHROM"], entry["POS"], entry["REF"], entry["ALT"],
                    entry["consequence"], entry["impact"], entry["shared_n"]):
            ws.cell(row=r_idx, column=col, value=val)
            col += 1

        for ctc in ctc_ids:
            s = entry["states"][ctc]
            state_cell = ws.cell(row=r_idx, column=col, value=CTC_STATE_LABEL[s["state"]])
            style_cell(state_cell, CTC_STATE_STYLE[s["state"]])
            ws.cell(row=r_idx, column=col + 1, value=s["dp"])
            ws.cell(row=r_idx, column=col + 2, value=s["alt"])
            col += 3

        wbc = entry["wbc"]
        tier_cell = ws.cell(row=r_idx, column=col, value=wbc["tier"])
        style_cell(tier_cell, WBC_TIER_STYLE.get(wbc["tier"], "neutral"), bold=(wbc["tier"] == "strong_signal"))
        ws.cell(row=r_idx, column=col + 1, value=wbc["gt"])
        ws.cell(row=r_idx, column=col + 2, value=wbc["dp"])
        ws.cell(row=r_idx, column=col + 3, value=wbc["alt"])

    autosize(ws, headers)


def write_shared_variant_matrix_tsv(path, computed_rows, ctc_ids):
    """Wide TSV for the SNV report's shared-variant heatmap: one row per shared
    variant, one {ctc}_state column per CTC, plus the WBC tier block. Same
    rows/states as the Excel sheet, minus per-cell DP/alt (the heatmap only
    needs the categorical state)."""
    fieldnames = ["var_id", "gene_symbol", "CHROM", "POS", "REF", "ALT", "shared_n"]
    fieldnames += [f"{c}_state" for c in ctc_ids]
    fieldnames += ["WBC_tier", "WBC_GT", "WBC_DP", "WBC_alt"]

    with open(path, "w", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for entry in computed_rows:
            out_row = {
                "var_id": entry["var_id"],
                "gene_symbol": entry["gene_symbol"],
                "CHROM": entry["CHROM"],
                "POS": entry["POS"],
                "REF": entry["REF"],
                "ALT": entry["ALT"],
                "shared_n": entry["shared_n"],
            }
            for ctc in ctc_ids:
                out_row[f"{ctc}_state"] = entry["states"][ctc]["state"]
            wbc = entry["wbc"]
            out_row["WBC_tier"] = wbc["tier"]
            out_row["WBC_GT"] = wbc["gt"]
            out_row["WBC_DP"] = wbc["dp"]
            out_row["WBC_alt"] = wbc["alt"]
            writer.writerow(out_row)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--patient-id", required=True)
    parser.add_argument("--long-table", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--matrix-out", required=True)
    parser.add_argument("--wbc-split-vcf", default="NA")
    parser.add_argument("ctc_split_vcfs", nargs="+")
    args = parser.parse_args()

    rows = load_long_table(args.long_table)
    ctc_ids = sorted({row["ctc_id"] for row in rows}, key=natural_sort_key)
    split_vcf_by_cell = map_split_vcfs_by_cell(args.ctc_split_vcfs)

    all_keys = {(row["CHROM"], row["POS"], row["REF"], row["ALT"]) for row in rows}
    wbc_path = args.wbc_split_vcf if args.wbc_split_vcf and args.wbc_split_vcf != "NA" else None
    wbc_raw = load_raw_records(wbc_path, all_keys)

    shared_computed = compute_shared_variant_states(rows, ctc_ids, split_vcf_by_cell, wbc_raw)

    wb = Workbook()
    wb.remove(wb.active)
    build_all_variants_sheet(wb, rows, wbc_raw)
    build_shared_variants_sheet(wb, shared_computed, ctc_ids)
    wb.save(args.out)
    write_shared_variant_matrix_tsv(args.matrix_out, shared_computed, ctc_ids)
    print(f"Wrote {args.out}: {len(rows)} final-call rows, {len({r['var_id'] for r in rows})} distinct variants")
    print(f"Wrote {args.matrix_out}: {len(shared_computed)} shared variants")


if __name__ == "__main__":
    main()
